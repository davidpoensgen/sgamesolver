"""Console script to run timings.

Specification - which homotopy, which game sizes, how many runs, which parameters for games/homotopy/solver etc -
are to be given in the excel files.

Usage:
1) To run an existing file (or multiple files):
python excel_timings.py filename0 [filename1 filename2 ....]
-> filenames can include folders; can omit .xlsx extension.
-> add flag -SD to shutdown once all files are done (on windows).

2) To create xlsx-files instead, just add flag -m:
python excel_timings.py -m filename0 [filename1 filename2 ....]
(make sure to adapt the new files and then run them as above)
-> Alternatively, just copy an existing file, make any desired changes,
   and delete all rows (except header) from table "runs".

3) To update the summary sheet of the specified files (instead of running them), use flag -s.
(The summary is updated upon completion, but it might be necessary to do so manually after errors.)
"""
import numpy as np

import sgamesolver
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
from socket import gethostname
import os
import argparse
import shutil

HOMOTOPIES = {
    "QRE": sgamesolver.homotopy.QRE,
    "LogTracing": sgamesolver.homotopy.LogTracing,
}


def make_file(filename):
    if filename[-5:] != ".xlsx":
        filename = filename + ".xlsx"
    print('~' * 75)

    if os.path.isfile(filename):
        answer = input(f'"{filename}" already exists. \n\n'
                       f'Type "yes" to OVERWRITE ALL file contents, anything else to cancel: \n')
        if answer != 'yes':
            print('Creating file canceled.')
            print('~' * 75)
            return
        else:
            print("\n")

    wb = openpyxl.Workbook()
    spec = wb.active
    spec.title = "Specification"

    spec.append(["homotopy", f"{'/'.join([h for h in HOMOTOPIES])}"])
    spec.append(["computer", ""])
    spec.append(["solver parameters", ""])
    spec.append(["homotopy parameters", ""])
    spec.append(["", ""])
    spec.append(["description", ""])
    spec.append(["file created on", datetime.now().strftime("%Y-%m-%d, %H:%M:%S")])
    spec.column_dimensions["A"].width = 22
    spec.column_dimensions["B"].width = 240

    games = wb.create_sheet("Games")
    games.append(["S", "I", "A", "count", "solver parameters", "homotopy parameters", "game parameters"])
    games.column_dimensions["E"].width = 32
    games.column_dimensions["F"].width = 32
    games.column_dimensions["G"].width = 32

    for S in [1, 2, 5, 10, 20, 50, 100, 200, 400]:
        for I in [2, 3, 4, 5]:
            for A in [2, 4, 8]:
                games.append([S, I, A, 100])

    runs = wb.create_sheet("Runs")
    runs.append(
        ["S", "I", "A", "number", "seed", "success", "steps", "time (s)",
         "failure reason", "exception", "date", "machine"])
    runs.column_dimensions["H"].width = 15
    runs.column_dimensions["I"].width = 20
    runs.column_dimensions["J"].width = 20
    runs.column_dimensions["K"].width = 20
    runs.column_dimensions["L"].width = 20

    summary = wb.create_sheet("Summary")
    summary.append(["S", "I", "A", "total runs", "successful", "success %",
                    "av time", "av time (s)", "std time (s)", "av steps", "std steps"])

    wb.save(filename=filename)
    print(f'"{filename}" created.\n'
          f'-> Remember to choose a homotopy and adjust parameters and run counts.')
    print('~' * 75)


def run_file(filename):
    if filename[-5:] != ".xlsx":
        filename = filename + ".xlsx"
    wb = openpyxl.load_workbook(filename=filename)

    def save_file():
        # make a backup copy before even attempting to save - safety first!
        shutil.copy(filename, filename + ".backup")
        try:
            wb.save(filename=filename)
            nonlocal time_saved
            time_saved = datetime.now()
            print('~' * 75)
            print(f'{datetime.now().strftime("%H:%M:%S")} > {filename} saved. > Total running time: '
                  f'{str(datetime.now() - time_started).split(".")[0]}')
            print('~' * 75)
        except PermissionError:
            print('~' * 75)
            input(f"Cannot write to {filename}. Please close excel and press ENTER to try again.")
            save_file()
        # KeyboardInterrupt while saving sometimes corrupts file - so I'll just catch it to be safe.
        # (user should just try again after saving)
        except KeyboardInterrupt:
            pass

    time_started = datetime.now()
    time_saved = datetime.now()
    save_interval_seconds = 300

    spec = wb["Specification"]
    homotopy_string = spec['B1'].value
    if homotopy_string not in HOMOTOPIES:
        print('~' * 75)
        if homotopy_string is not None:
            print(f'ERROR: Homotopy "{homotopy_string}" given in the excel file, does not exist.')
        else:
            print('ERROR: No homotopy specified in the excel file.')
        print(f'Currently available homotopies are: {", ".join([h for h in HOMOTOPIES])}\n'
              'Please adapt the file and run again.\n'
              '~' * 75)
        return

    homotopy_constructor = HOMOTOPIES[homotopy_string]
    solver_parameters_all = str_to_dict(spec['B3'].value or "")
    homotopy_parameters_all = str_to_dict(spec['B4'].value or "")

    runs = wb["Runs"]

    games_pd = pd.read_excel(filename, sheet_name="Games", keep_default_na=False)
    runs_pd = pd.read_excel(filename, sheet_name="Runs")

    machine = gethostname()
    print('~' * 75)
    print(f'{datetime.now().strftime("%H:%M:%S")} > Starting to run {filename}.')
    print('~' * 75)

    for idx, game_spec in games_pd.iterrows():
        S, I, A, count, solver_parameters, homotopy_parameters, game_parameters = game_spec
        solver_parameters = str_to_dict(solver_parameters or "")
        homotopy_parameters = str_to_dict(homotopy_parameters or "")
        game_parameters = str_to_dict(game_parameters or "")

        numbers_done = list(runs_pd.query(f'S=={S} & I=={I} & A=={A}')['number'])
        success_count = runs_pd.query(f'S=={S} & I=={I} & A=={A} & success == True').shape[0]
        for number in range(100_000):
            if success_count >= count:
                break
            if number in numbers_done:
                continue

            try:
                seed = int(f'{S:02d}{I:02d}{A:02d}{number:04d}')
                print(f'{datetime.now().strftime("%H:%M:%S")} > {S}-{I}-{A}: {number:2d} ', end="", flush=True)
                date = datetime.now().strftime("%Y-%m-%d, %H:%M:%S")

                game = sgamesolver.SGame.random_game(S, I, A, seed=seed, **game_parameters)

                # if the same parameter is specified both in _parameters and parameters_all, the former will override.
                homotopy = homotopy_constructor(game, **{**homotopy_parameters_all, **homotopy_parameters})
                homotopy.solver_setup()
                homotopy.solver.verbose = 0
                homotopy.solver.set_parameters(**{**solver_parameters_all, **solver_parameters})
            except KeyboardInterrupt:
                print('\n' + '+!+' * 25)
                print(f'{datetime.now().strftime("%H:%M:%S")} > '
                      f'KEYBOARD INTERRUPT > PLEASE WAIT WHILE SAVING {filename}.')
                save_file()
                summarize_file(filename)
                raise

            try:
                result = homotopy.solver.start()
            except Exception as exception:
                runs.append([S, I, A, number, seed, False, homotopy.solver.step, "", "", str(exception), date, machine])
                print("Exception: ", str(exception))
                if (datetime.now() - time_saved).total_seconds() >= save_interval_seconds:
                    save_file()
            except KeyboardInterrupt:
                print('\n' + '+!+' * 25)
                print(f'{datetime.now().strftime("%H:%M:%S")} > '
                      f'KEYBOARD INTERRUPT > PLEASE WAIT WHILE SAVING {filename}.')
                save_file()
                summarize_file(filename)
                raise
            else:
                runs.append([S, I, A, number, seed, result["success"], result["steps"],
                             result["time"], result['failure reason'] or "", "", date, machine])
                if result["success"]:
                    success_count += 1
                    print(f"> {result['time']:.1f}s, {result['steps']} steps")
                else:
                    print(f"> {result['time']:.1f}s, {result['steps']} steps. Failed: {result['failure reason']}")

                if (datetime.now() - time_saved).total_seconds() >= save_interval_seconds:
                    save_file()

            del game
            del homotopy

    save_file()
    summarize_file(filename)
    print(f'{datetime.now().strftime("%H:%M:%S")} > DONE!')
    print('~' * 75)


def summarize_file(filename):
    if filename[-5:] != ".xlsx":
        filename = filename + ".xlsx"
    wb = openpyxl.load_workbook(filename=filename)

    summary = wb["Summary"]
    summary.delete_rows(2, 500)
    games_pd = pd.read_excel(filename, sheet_name="Games", keep_default_na=False)
    # only unique S-I-A-combinations are needed:
    games_pd.drop_duplicates(subset=['S', 'I', 'A'], keep='first', inplace=True)
    runs_pd = pd.read_excel(filename, sheet_name="Runs")

    for idx, game_spec in games_pd.iterrows():
        S, I, A, count, solver_parameters, homotopy_parameters, game_parameters = game_spec
        all_runs = runs_pd.query(f'S=={S} & I=={I} & A=={A}')
        successful_runs = runs_pd.query(f'S=={S} & I=={I} & A=={A} & success == True')
        if all_runs.shape[0] == 0:
            continue
        if successful_runs.shape[0] == 0:
            summary.append([S, I, A, all_runs.shape[0], 0, 0])
            continue

        summary.append([
            S, I, A,
            all_runs.shape[0],
            successful_runs.shape[0],
            round(successful_runs.shape[0] / all_runs.shape[0] * 100, 1),
            str(timedelta(seconds=round(successful_runs["time (s)"].mean()))),
            round(successful_runs["time (s)"].mean(), 2),
            round(successful_runs["time (s)"].std(), 2),
            round(successful_runs["steps"].mean(), 1),
            round(successful_runs["steps"].std(), 1)
        ])

    wb.save(filename)
    print(f'Summary of "{filename}" updated.')


def latex_file(filename):
    if filename[-5:] != '.xlsx':
        filename = filename + '.xlsx'
    summary_pd = pd.read_excel(filename, sheet_name='Summary')
    S_counts = sorted(summary_pd['S'].unique().tolist())
    I_counts = sorted(summary_pd['I'].unique().tolist())
    latex = '\\documentclass{article}\n\\usepackage{booktabs}\n\\usepackage{multirow}\n\n\\begin{document}\n'
    latex += '\\begin{tabular}{r@{\\hskip .4cm}r@{\\hskip .6cm}' \
             + 'r@{\\hskip .15cm}l@{\hskip .1cm}' * len(I_counts) + '}\n\\toprule\n'
    latex += f'\\multirow{{2}}{{*}}{{$|S|$}} & \\multirow{{2}}{{*}}{{$|A|$}} & ' \
             f'\\multicolumn{{{2 * len(I_counts)}}}{{c}}{{$|I|$}} \\\\ \\cmidrule(lr){{3-{2 + 2 * len(I_counts)}}} \n'
    latex += ' & & ' + ' & '.join(f'\\multicolumn{{2}}{{c}}{{{I}}}' for I in I_counts) + '\\\\ \\midrule \n '
    # row of \hphantoms -> for equal spacing of columns
    latex += ' & & ' + ' & '.join('\\hphantom{1:00:00} & \\hphantom{\\footnotesize{1:00:00}}' for _ in I_counts) \
             + '\\\\[-4pt] \n '
    for S in S_counts:
        # get all |A| for the current |S|:
        A_counts = sorted(summary_pd.query(f'S=={S}')['A'].unique().tolist())
        for A in A_counts:
            # |S| only on first line of each block. phantom makes alignment look nicer:
            if A == A_counts[0]:
                latex += f'${S}\\phantom{{|}}$ & '
            else:
                latex += ' & '
            latex += f'${A}\\phantom{{|}}$ & '
            for I in I_counts:
                pd_row = summary_pd.query(f'S=={S} & I=={I} & A=={A}')
                if len(pd_row) > 1:
                    raise ValueError(f'Summary contains multiple rows for S,I,A = {S},{I},{A}.')
                # skip cell if no row exists, or row exists but only failed runs (-> avg time not defined)
                elif len(pd_row) == 0 or pd_row.head(1)['successful'].item() == 0:
                    latex += '&'
                elif len(pd_row) == 1:
                    avg_s = pd_row.head(1)['av time (s)'].item()
                    std_s = pd_row.head(1)['std time (s)'].item()
                    latex += f'{time_format(avg_s)} & \\footnotesize{{{time_format(std_s)}}}'
                    # alternative representation - secs w/ 2 decimals::
                    # latex += f'{avg_s} & \\footnotesize{{{std_s}}}'
                # last column will end in \\ rather than &:
                if I != I_counts[-1]:
                    latex += ' & '
            # end S-A-row:
            latex += ' \\\\ \n'
        # end S-block:
        latex += '&' * (len(I_counts) * 2 + 1) + '\\\\[-4pt] \n'
    # end table and document:
    latex += '\\bottomrule \n'
    latex += '\\end{tabular}\n'
    latex += '\\end{document}'

    tex_filename = filename[:-5] + '.tex'
    with open(tex_filename, 'w') as tex_file:
        tex_file.write(latex)
        print(f'Latex file "{tex_filename}" created or updated.')


def time_format(secs):
    """Format secs as m:ss / h:mm:ss"""
    # rarely, secs might be np.nan (e.g. std dev. after a single run) - in this case, just return a dash
    if np.isnan(secs):
        return "--"
    minutes, seconds = divmod(round(secs), 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        string = f'{hours}:{minutes:02}:{seconds:02}'
    elif minutes:
        string = f'{minutes}:{seconds:02}'
    else:
        string = f'0:{seconds:02}'
    return string


def str_to_dict(str_):
    out = {}
    parameters = str_.split(",")
    for parameter in parameters:
        if parameter and parameter.strip():  # skip empty string, just whitespace etc.
            key, value = parameter.split('=')
            value = value.strip()
            try:  # cast to float:
                value = float(value)
                # convert to int if sensible:
                if value == int(value):
                    value = int(value)
            except ValueError:
                # convert some special strings:
                if value == "True":
                    value = True
                elif value == "False":
                    value = False
                if value == "None":
                    value = None
                # all other string are preserved
            out[key.strip()] = value
    return out


if __name__ == '__main__':
    import sys

    parser = argparse.ArgumentParser(description='Run (or create, summarize) a timings file.')
    parser.add_argument('filenames', metavar='filename', nargs='+',
                        help='File(s) to be run (may include path; may omit file extension .xslx)')
    group = parser.add_mutually_exclusive_group()
    group.add_argument('-m', action='store_true', help='Create the file(s) instead of running it.')
    group.add_argument('-s', action='store_true', help='Summarize the file(s) instead of running it.')
    group.add_argument('-l', action='store_true', help='Summarize the file(s) and create a latex table for each. '
                                                       '(As .tex file; will overwrite without prompt.)')
    group.add_argument('-SD', action='store_true', help='Run and shutdown computer once all files are done.')

    args = parser.parse_args()

    if args.m:
        for file in args.filenames:
            make_file(file)
        sys.exit()

    if args.s:
        for file in args.filenames:
            if file[-5:] != ".xlsx":
                file = file + ".xlsx"
            if not os.path.isfile(file):
                print(f'ERROR: "{file}" not found.')
            else:
                summarize_file(file)
        sys.exit()

    if args.l:
        for file in args.filenames:
            if file[-5:] != ".xlsx":
                file = file + ".xlsx"
            if not os.path.isfile(file):
                print(f'ERROR: "{file}" not found.')
            else:
                summarize_file(file)
                latex_file(file)
        sys.exit()

    # first pass: check if all files can be found:
    missing_file = False
    for file in args.filenames:
        if file[-5:] != ".xlsx":
            file = file + ".xlsx"
        if not os.path.isfile(file):
            missing_file = True
            print(f'ERROR: "{file}" not found.')
    if missing_file:
        sys.exit()
    # second pass: actually (attempt to) run all files.
    try:
        for file in args.filenames:
            run_file(file)
        if args.SD:
            import subprocess

            subprocess.run(["shutdown", "-s"])
    except Exception:
        # any exception the running code does not catch (besides KeyboardInterrupt).
        # (e.g. illegal kwargs when setting up homotopies or similar)
        # this is to ensure shutdown goes through even if something unexpected happens.
        if args.SD:
            import subprocess

            subprocess.run(["shutdown", "-s"])
        else:
            raise
